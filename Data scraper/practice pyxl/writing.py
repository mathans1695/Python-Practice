from openpyxl import load_workbook
import csv

class Company:
	def __init__(self, last_name, company, email, industry, city):
		self.last_name = last_name
		self.company = company
		self.email = email
		self.industry = industry
		self.city = city

workbook = load_workbook('Tamilnadu_companies_data.xlsx')
sheetnames = workbook.sheetnames

spreadsheet = workbook[sheetnames[len(sheetnames)-1]]

listOfCompanies = {}

for row in spreadsheet.iter_rows(min_row=2,
								min_col=1,
								max_col=5,
								values_only=True
								):
	listOfCompanies[row[1]] = Company(row[0], row[1], row[2], row[3], row[4])
	
with open('trial_data.csv', 'w', newline='') as csvfile:
	writer = csv.writer(csvfile)
	
	writer.writerow(['Last Name', 'Company', 'Email', 'Industry', 'City'])
	
	for company in listOfCompanies.values():
		writer.writerow([company.last_name, company.company, company.email, company.industry, company.city])


